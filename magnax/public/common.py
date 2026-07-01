import asyncio
import json
import os
import platform
import re
import shutil
import time
from collections import OrderedDict
import requests
from loguru import logger
from tqdm import tqdm
import socket
from urllib.request import urlopen
import ssl
import openpyxl
import psutil
import signal
import cv2
from functools import wraps
from jinja2 import Environment, FileSystemLoader

# 使用 pymobiledevice3 进行 iOS 设备控制
try:
    from pymobiledevice3.lockdown import LockdownClient, create_using_usbmux
    from pymobiledevice3.usbmux import list_devices as pmd3_list_devices
    PMD3_AVAILABLE = True
except ImportError:
    LockdownClient = None
    create_using_usbmux = None
    pmd3_list_devices = None
    PMD3_AVAILABLE = False
from magnax.public.adb import adb


# 生成静态 HTML 报告时图表的最大数据点数（与前端动态图表默认值保持一致）。
# 超过该值会用 LTTB 降采样，避免长时间采集后报告体积与渲染时间随采集时长线性膨胀。
REPORT_CHART_MAX_POINTS = 1000


def downsample_lttb(data: list, target_points: int) -> list:
    """
    LTTB (Largest Triangle Three Buckets) 降采样算法
    专为时序数据设计，保留视觉特征（峰值、谷值、趋势）

    Args:
        data: [{"x": timestamp, "y": value}, ...] 格式的时序数据
        target_points: 目标数据点数量（建议 500-2000）

    Returns:
        降采样后的数据列表，保留首尾点和关键特征点
    """
    n = len(data)
    if n <= target_points or target_points < 3:
        return data
    # 复用 _lttb_indices(已对末点重复做去重),避免两份 LTTB 实现各自维护、各自出 bug
    return [data[i] for i in _lttb_indices(data, target_points)]


def _lttb_indices(data: list, target_points: int) -> list:
    """返回 LTTB 选中的原始索引列表(供多 series 对齐降采样复用)。"""
    n = len(data)
    if n <= target_points or target_points < 3:
        return list(range(n))
    indices = [0]
    bucket_size = (n - 2) / (target_points - 2)
    a = 0
    for i in range(target_points - 2):
        bucket_start = int((i + 1) * bucket_size) + 1
        bucket_end = min(int((i + 2) * bucket_size) + 1, n - 1)
        next_start = int((i + 2) * bucket_size) + 1
        next_end = min(int((i + 3) * bucket_size) + 1, n)
        if next_end > next_start:
            avg_x = sum(j for j in range(next_start, next_end)) / (next_end - next_start)
            avg_y = sum(data[j]['y'] for j in range(next_start, next_end)) / (next_end - next_start)
        else:
            avg_x = next_start
            avg_y = data[min(next_start, n - 1)]['y']
        max_area = -1
        max_idx = bucket_start
        for j in range(bucket_start, bucket_end):
            area = abs(
                (a - avg_x) * (data[j]['y'] - data[a]['y']) -
                (a - j) * (avg_y - data[a]['y'])
            )
            if area > max_area:
                max_area = area
                max_idx = j
        indices.append(max_idx)
        a = max_idx
    indices.append(n - 1)
    # 末桶 bucket_start 恒为 n-1(因 (target_points-2)*bucket_size == n-2),空桶使 max_idx 默认取 n-1,
    # 循环内 append 一次、循环后再 append(n-1) 一次 -> 最后一个点被选两次 -> x 重复 ->
    # ApexCharts category 轴整条序列渲染失败(长时间报告统计图全空)。按序去重修复。
    seen = set()
    uniq = []
    for idx in indices:
        if idx not in seen:
            seen.add(idx)
            uniq.append(idx)
    return uniq


def downsample_lttb_multi(series_list: list, target_points: int) -> list:
    """对多个等长的同源 series 用统一采样索引降采样。

    多条线来自同一次采集(x 时间戳相同),但若各自独立 LTTB 选点会导致 x 不对齐,
    多线图(如 iOS 电池的温度/电流/电压/功率)将因坐标错位而画不出线。
    这里以第一条非空 series 计算 LTTB 索引,应用到所有等长 series,保证 x 对齐。
    """
    base = next((s for s in series_list if s), None)
    if base is None:
        return series_list
    n = len(base)
    if target_points and 0 < target_points < n and target_points >= 3:
        idx = _lttb_indices(base, target_points)
    else:
        idx = list(range(n))
    base_x = [base[i]['x'] for i in idx]
    return [
        [{'x': base_x[k], 'y': s[idx[k]]['y']} for k in range(len(idx))]
        if len(s) == n else s
        for s in series_list
    ]


def get_ios_lockdown_client_in_common(device_id):
    """获取iOS设备的lockdown client (common.py专用版本)"""
    try:
        if create_using_usbmux is None:
            logger.warning("pymobiledevice3 not available, some iOS features may not work")
            return None
        
        return create_using_usbmux(serial=device_id)
    except Exception as e:
        logger.error(f"Failed to create lockdown client for device {device_id}: {e}")
        return None

def _pmd3_list_devices_sync():
    """兼容新旧 pymobiledevice3:新版 list_devices 可能返回协程,需 await 取结果。"""
    devices = pmd3_list_devices()
    if asyncio.iscoroutine(devices):
        devices = asyncio.run(devices)
    return devices

def get_ios_device_udid_list():
    """获取连接的iOS设备UDID列表"""
    if not PMD3_AVAILABLE:
        logger.warning("pymobiledevice3 not available")
        return []
    try:
        devices = _pmd3_list_devices_sync()
        return [device.serial for device in devices]
    except Exception as e:
        logger.error(f"Failed to get iOS device list: {e}")
        return []

class Platform:
    Android = 'Android'
    iOS = 'iOS'
    Mac = 'MacOS'
    Windows = 'Windows'

class Devices:

    def __init__(self, platform=Platform.Android):
        self.platform = platform
        self.adb = adb.adb_path

    def execCmd(self, cmd):
        """Execute the command to get the terminal print result"""
        r = os.popen(cmd)
        try:
            text = r.buffer.read().decode(encoding='gbk').replace('\x1b[0m','').strip()
        except UnicodeDecodeError:
            text = r.buffer.read().decode(encoding='utf-8').replace('\x1b[0m','').strip()
        finally:
            r.close()
        return text

    def filterType(self):
        """Select the pipe filtering method according to the system"""
        filtertype = ('grep', 'findstr')[platform.system() == Platform.Windows]
        return filtertype

    def getDeviceIds(self):
        """Get all connected device ids"""
        Ids = list(os.popen(f"{self.adb} devices").readlines())
        deviceIds = []
        for i in range(1, len(Ids) - 1):
            id, state = Ids[i].strip().split()
            if state == 'device':
                deviceIds.append(id)
        return deviceIds

    def getDevicesName(self, deviceId):
        """Get the device name of the Android corresponding device ID"""
        try:
            lines = os.popen(f'{self.adb} -s {deviceId} shell getprop ro.product.model').readlines()
            devices_name = lines[0].strip() if lines else deviceId
        except Exception:
            devices_name = deviceId
        return devices_name

    def getDevices(self):
        """Get all Android devices"""
        DeviceIds = self.getDeviceIds()
        Devices = [f'{id}({self.getDevicesName(id)})' for id in DeviceIds]
        logger.info('Connected devices: {}'.format(Devices))
        return Devices

    def getIdbyDevice(self, deviceinfo, platform):
        """Obtain the corresponding device id according to the Android device information"""
        if platform == Platform.Android:
            deviceId = re.sub(u"\\(.*?\\)|\\{.*?}|\\[.*?]", "", deviceinfo)
            if deviceId not in self.getDeviceIds():
                raise Exception('no device found')
        else:
            deviceId = deviceinfo
        return deviceId
    
    def getSdkVersion(self, deviceId):
        version = adb.shell(cmd='getprop ro.build.version.sdk', deviceId=deviceId)
        return version
    
    def getCpuCores(self, deviceId):
        """get Android cpu cores"""
        cmd = 'cat /sys/devices/system/cpu/online'
        result = adb.shell(cmd=cmd, deviceId=deviceId)
        try:
            # online 可能是 "0"、"0-3"、"0,2-7" 等形式,取最大核编号 + 1
            nums = max(int(n) for n in re.findall(r'\d+', result)) + 1
        except Exception:
            nums = 1
        return nums

    def getPid(self, deviceId, pkgName):
        """Get the pid corresponding to the Android package name"""
        try:
            sdkversion = self.getSdkVersion(deviceId)
            if sdkversion and int(sdkversion) < 26:
                result = os.popen(f"{self.adb} -s {deviceId} shell ps | {self.filterType()} {pkgName}").readlines()
                processList = ['{}:{}'.format(process.split()[1],process.split()[8]) for process in result]
            else:
                result = os.popen(f"{self.adb} -s {deviceId} shell ps -ef | {self.filterType()} {pkgName}").readlines()
                processList = ['{}:{}'.format(process.split()[1],process.split()[7]) for process in result]
            for i in range(len(processList)):
                if processList[i].count(':') == 1:
                    index = processList.index(processList[i])
                    processList.insert(0, processList.pop(index))
                    break
            if len(processList) == 0:
               logger.warning('{}: no pid found'.format(pkgName))     
        except Exception as e:
            processList = []
            logger.exception(e)
        return processList

    def checkPkgname(self, pkgname):
        flag = True
        replace_list = ['com.google']
        for i in replace_list:
            if i in pkgname:
                flag = False
        return flag

    def getPkgname(self, deviceId):
        """Get all package names of Android devices"""
        pkginfo = os.popen(f"{self.adb} -s {deviceId} shell pm list packages --user 0")
        pkglist = [p.lstrip('package').lstrip(":").strip() for p in pkginfo]
        if pkglist.__len__() > 0:
            return pkglist
        else:
            pkginfo = os.popen(f"{self.adb} -s {deviceId} shell pm list packages")
            pkglist = [p.lstrip('package').lstrip(":").strip() for p in pkginfo]
            return pkglist

    def getDeviceInfoByiOS(self):
        """Get a list of all successfully connected iOS devices"""
        deviceInfo = get_ios_device_udid_list()
        logger.info('Connected devices: {}'.format(deviceInfo))    
        return deviceInfo

    def getPkgnameByiOS(self, udid):
        """Get all package names of the corresponding iOS device"""
        try:
            lockdown_client = get_ios_lockdown_client_in_common(udid)
            if lockdown_client is None:
                logger.error("Failed to get lockdown client for iOS package list")
                return []
            
            from pymobiledevice3.services.installation_proxy import InstallationProxyService
            installation = InstallationProxyService(lockdown=lockdown_client)

            # 获取用户安装的应用列表（新版 API 返回字典，key 为 bundle identifier）
            apps = installation.get_apps(application_type='User')
            pkgNames = list(apps.keys())
            return pkgNames
        except Exception as e:
            logger.error(f"Failed to get iOS package names: {e}")
            return []
    
    def get_pc_ip(self):
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(('8.8.8.8', 80))
            ip = s.getsockname()[0]
        except Exception:
            logger.error('get local ip failed')
            ip = '127.0.0.1'
        finally:
            s.close()
        return ip
    
    def get_device_ip(self, deviceId):
        content = os.popen(f"{self.adb} -s {deviceId} shell ip addr show wlan0").read()
        logger.info(content)
        math_obj = re.search(r'inet\s(\d+\.\d+\.\d+\.\d+).*?wlan0', content)
        if math_obj and math_obj.group(1):
            return math_obj.group(1)
        return None
    
    def devicesCheck(self, platform, deviceid=None, pkgname=None):
        """Check the device environment"""
        match(platform):
            case Platform.Android:
                if len(self.getDeviceIds()) == 0:
                    raise Exception('no devices found')
                if len(self.getPid(deviceId=deviceid, pkgName=pkgname)) == 0:
                    raise Exception('no process found')
            case Platform.iOS:
                if len(self.getDeviceInfoByiOS()) == 0:
                    raise Exception('no devices found')
            case _:
                raise Exception('platform must be Android or iOS')        
            
    def getDdeviceDetail(self, deviceId, platform):
        result = dict()
        match(platform):
            case Platform.Android:
                result['brand'] = adb.shell(cmd='getprop ro.product.brand', deviceId=deviceId)
                result['name'] = adb.shell(cmd='getprop ro.product.model', deviceId=deviceId)
                result['version'] = adb.shell(cmd='getprop ro.build.version.release', deviceId=deviceId)
                result['serialno'] = adb.shell(cmd='getprop ro.serialno', deviceId=deviceId)
                cmd = f'ip addr show wlan0 | {self.filterType()} link/ether'
                wifiadr_content = adb.shell(cmd=cmd, deviceId=deviceId)                
                result['wifiadr'] = Method._index(wifiadr_content.split(), 1, '')
                result['cpu_cores'] = self.getCpuCores(deviceId)
                result['physical_size'] = adb.shell(cmd='wm size', deviceId=deviceId).replace('Physical size:','').strip()
            case Platform.iOS:
                try:
                    lockdown_client = get_ios_lockdown_client_in_common(deviceId)
                    if lockdown_client is None:
                        logger.error("Failed to get lockdown client for iOS device details")
                        return {'brand': '', 'name': '', 'version': '', 'serialno': deviceId, 'wifiadr': '', 'cpu_cores': 0, 'physical_size': ''}
                    
                    # 从lockdown client获取设备信息
                    device_info = lockdown_client.all_values
                    
                    result['brand'] = device_info.get("DeviceClass", "")
                    result['name'] = device_info.get("DeviceName", "")
                    result['version'] = device_info.get("ProductVersion", "")
                    result['serialno'] = deviceId
                    result['wifiadr'] = device_info.get("WiFiAddress", "")
                    result['cpu_cores'] = 0
                    result['physical_size'] = self.getPhysicalSzieOfiOS(deviceId)
                except Exception as e:
                    logger.error(f"Failed to get iOS device details: {e}")
                    result = {'brand': '', 'name': '', 'version': '', 'serialno': deviceId, 'wifiadr': '', 'cpu_cores': 0, 'physical_size': ''}
            case _:
                raise Exception('{} is undefined'.format(platform)) 
        return result
    
    def getPhysicalSzieOfiOS(self, deviceId):
        try:
            lockdown_client = get_ios_lockdown_client_in_common(deviceId)
            if lockdown_client is None:
                logger.error("Failed to get lockdown client for iOS screen info")
                return ''
            
            # 获取屏幕信息
            device_info = lockdown_client.all_values
            screen_width = device_info.get('ScreenWidth', 0)
            screen_height = device_info.get('ScreenHeight', 0)
            
            if screen_width and screen_height:
                PhysicalSzie = '{}x{}'.format(screen_width, screen_height)
            else:
                PhysicalSzie = ''
        except Exception as e:
            PhysicalSzie = ''  
            logger.exception(e)  
        return PhysicalSzie
    
    def getCurrentActivity(self, deviceId):
        result = adb.shell(cmd='dumpsys window | {} mCurrentFocus'.format(self.filterType()), deviceId=deviceId)
        if result.__contains__('mCurrentFocus'):
            activity = str(result).split(' ')[-1].replace('}','') 
            return activity
        else:
            raise Exception('No activity found')

    def getStartupTimeByAndroid(self, activity, deviceId):
        result = adb.shell(cmd='am start -W {}'.format(activity), deviceId=deviceId)
        return result

    def getStartupTimeByiOS(self, pkgname):
        try:
            import ios_device
        except ImportError:
            logger.error('py-ios-devices not found, please run [pip install py-ios-devices]') 
        result = self.execCmd('pyidevice instruments app_lifecycle -b {}'.format(pkgname))       
        return result          

class File:

    def __init__(self, fileroot='.'):
        self.fileroot = fileroot
        self.report_dir = self.get_repordir()
        # log 文件全量解析缓存：(scene, filename) -> (mtime, log_data_list, target_data_list)
        # 同一份 log 在均值计算与图表生成中会被多次读取，缓存可避免重复读盘与解析。
        self._log_cache = OrderedDict()
        self._log_cache_limit = 32

    def _safe_remove_file(self, filepath, max_retries=5, retry_delay=1):
        """安全删除文件，处理文件被占用的情况"""
        for attempt in range(max_retries):
            try:
                if os.path.exists(filepath):
                    os.remove(filepath)
                    logger.info(f'文件删除成功: {filepath}')
                return True
                
            except PermissionError as e:
                if attempt < max_retries - 1:
                    logger.warning(f'文件被占用，等待后重试删除 ({attempt + 1}/{max_retries}): {filepath}')
                    time.sleep(retry_delay)
                    retry_delay *= 2  # 指数退避
                else:
                    logger.error(f'删除文件失败，已达到最大重试次数: {filepath}')
                    return False
            except Exception as e:
                logger.warning(f'删除文件时发生错误: {filepath}, 错误: {e}')
                return False
        return False

    def clear_file(self):
        logger.info('Clean up useless files ...')
        if os.path.exists(self.report_dir):
            files_to_remove = []
            for f in os.listdir(self.report_dir):
                filename = os.path.join(self.report_dir, f)
                if f.split(".")[-1] in ['log', 'json', 'mkv']:
                    files_to_remove.append(filename)
            
            # 先停止所有录屏进程，确保文件不被占用
            Scrcpy.stop_record()
            
            # 等待一些时间让进程完全结束
            time.sleep(2)
            
            # 安全删除文件
            for filename in files_to_remove:
                success = self._safe_remove_file(filename)
                if not success:
                    logger.warning(f'无法删除文件，将在下次清理时重试: {filename}')
                            
        logger.info('Clean up useless files success')

    def export_excel(self, platform, scene):
        logger.info('Exporting excel ...')
        android_log_file_list = ['cpu_app','cpu_sys','mem_total','mem_swap',
                                 'battery_level', 'battery_tem','upflow','downflow','fps','gpu']
        ios_log_file_list = ['cpu_app','cpu_sys', 'mem_total', 'battery_tem', 'battery_current', 
                             'battery_voltage', 'battery_power','upflow','downflow','fps','gpu']
        log_file_list = android_log_file_list if platform == 'Android' else ios_log_file_list
        wb = openpyxl.Workbook()
        # Remove the default sheet created by openpyxl
        wb.remove(wb.active)
        for name in log_file_list:
            ws = wb.create_sheet(title=name)
            ws.cell(row=1, column=1, value='Time')
            ws.cell(row=1, column=2, value='Value')
            row = 2  # start row (1-based, header is row 1)
            if os.path.exists(f'{self.report_dir}/{scene}/{name}.log'):
                with open(f'{self.report_dir}/{scene}/{name}.log', 'r', encoding='utf-8') as f:
                    for lines in f:
                        target = lines.split('=')
                        for i in range(len(target)):
                            ws.cell(row=row, column=i + 1, value=target[i])
                        row += 1
        xlsx_path = os.path.join(self.report_dir, scene, f'{scene}.xlsx')
        wb.save(xlsx_path)
        logger.info('Exporting excel success : {}'.format(xlsx_path))
        return xlsx_path   
    
    def make_android_html(self, scene, summary : dict, report_path=None):
        logger.info('Generating HTML ...')
        STATICPATH = os.path.dirname(os.path.realpath(__file__))
        file_loader = FileSystemLoader(os.path.join(STATICPATH, 'report_template'))
        env = Environment(loader=file_loader)
        template = env.get_template('android.html')
        if report_path:
            html_path = report_path
        else:
            html_path = os.path.join(self.report_dir, scene, 'report.html')   
        with open(html_path,'w+', encoding='utf-8') as fout:
            html_content = template.render(devices=summary['devices'],app=summary['app'],
                                           platform=summary['platform'],ctime=summary['ctime'],
                                           cpu_app=summary['cpu_app'],cpu_sys=summary['cpu_sys'],
                                           mem_total=summary['mem_total'],mem_swap=summary['mem_swap'],
                                           fps=summary['fps'],jank=summary['jank'],level=summary['level'],
                                           tem=summary['tem'],net_send=summary['net_send'],
                                           net_recv=summary['net_recv'],cpu_charts=summary['cpu_charts'],
                                           mem_charts=summary['mem_charts'],net_charts=summary['net_charts'],
                                           battery_charts=summary['battery_charts'],fps_charts=summary['fps_charts'],
                                           jank_charts=summary['jank_charts'],mem_detail_charts=summary['mem_detail_charts'],
                                           gpu=summary['gpu'], gpu_charts=summary['gpu_charts'])
            
            fout.write(html_content)
        logger.info('Generating HTML success : {}'.format(html_path))  
        return html_path
    
    def make_ios_html(self, scene, summary : dict, report_path=None):
        logger.info('Generating HTML ...')
        STATICPATH = os.path.dirname(os.path.realpath(__file__))
        file_loader = FileSystemLoader(os.path.join(STATICPATH, 'report_template'))
        env = Environment(loader=file_loader)
        template = env.get_template('ios.html')
        if report_path:
            html_path = report_path
        else:
            html_path = os.path.join(self.report_dir, scene, 'report.html')
        with open(html_path,'w+', encoding='utf-8') as fout:
            html_content = template.render(devices=summary['devices'],app=summary['app'],
                                           platform=summary['platform'],ctime=summary['ctime'],
                                           cpu_app=summary['cpu_app'],cpu_sys=summary['cpu_sys'],gpu=summary['gpu'],
                                           mem_total=summary['mem_total'],fps=summary['fps'],
                                           tem=summary['tem'],current=summary['current'],
                                           voltage=summary['voltage'],power=summary['power'],
                                           net_send=summary['net_send'],net_recv=summary['net_recv'],
                                           cpu_charts=summary['cpu_charts'],mem_charts=summary['mem_charts'],
                                           net_charts=summary['net_charts'],battery_charts=summary['battery_charts'],
                                           fps_charts=summary['fps_charts'],gpu_charts=summary['gpu_charts'])            
            fout.write(html_content)
        logger.info('Generating HTML success : {}'.format(html_path))  
        return html_path
  
    def filter_secen(self, scene):
        dirs = os.listdir(self.report_dir)
        dir_list = list(reversed(sorted(dirs, key=lambda x: os.path.getmtime(os.path.join(self.report_dir, x)))))
        if scene in dir_list:
            dir_list.remove(scene)
        return dir_list

    def get_repordir(self):
        # 报告固定存放在用户目录下的 ~/MagnaX/report,不再随启动时的当前工作目录
        # (os.getcwd())变化。此前用 os.getcwd()/report,从不同目录(python -m magnax
        # vs cd magnax && python debug.py)或双击启动时 cwd 不同,历史报告会分散在各自
        # 的 report/ 里,导致"二次启动加载不到历史报告"。固定到用户目录后从任何位置启动
        # 都指向同一处。
        report_dir = os.path.join(os.path.expanduser('~'), 'MagnaX', 'report')
        os.makedirs(report_dir, exist_ok=True)
        return report_dir

    def create_file(self, filename, content=''):
        if not os.path.exists(self.report_dir):
            os.mkdir(self.report_dir)
        with open(os.path.join(self.report_dir, filename), 'a+', encoding="utf-8") as file:
            file.write(content)

    def add_log(self, path, log_time, value):
        if value >= 0:
            with open(path, 'a+', encoding="utf-8") as file:
                file.write(f'{log_time}={str(value)}' + '\n')
    
    def record_net(self, type, send , recv):
        net_dict = dict()
        match(type):
            case 'pre':
                net_dict['send'] = send
                net_dict['recv'] = recv
                content = json.dumps(net_dict)
                self.create_file(filename='pre_net.json', content=content)
            case 'end':
                net_dict['send'] = send
                net_dict['recv'] = recv
                content = json.dumps(net_dict)
                self.create_file(filename='end_net.json', content=content)
            case _:
                logger.error('record network data failed')
    
    def _safe_move_file(self, src, dst, max_retries=5, retry_delay=1):
        """安全移动文件，处理文件被占用的情况"""
        for attempt in range(max_retries):
            try:
                # 检查目标路径，如果是目录则生成完整的目标文件路径
                if os.path.isdir(dst):
                    dst_file = os.path.join(dst, os.path.basename(src))
                else:
                    dst_file = dst
                
                # 如果目标文件已存在，先尝试删除（可能是之前不完整的移动）
                if os.path.exists(dst_file):
                    logger.warning(f'目标文件已存在，尝试删除: {dst_file}')
                    try:
                        os.remove(dst_file)
                        logger.info(f'成功删除已存在的目标文件: {dst_file}')
                    except Exception as delete_error:
                        logger.warning(f'无法删除已存在文件: {delete_error}')
                        # 如果无法删除，生成新的文件名
                        base, ext = os.path.splitext(dst_file)
                        counter = 1
                        while os.path.exists(dst_file):
                            dst_file = f"{base}_{counter}{ext}"
                            counter += 1
                        logger.info(f'生成新的目标文件名: {dst_file}')
                
                shutil.move(src, dst_file)
                logger.info(f'文件移动成功: {src} -> {dst_file}')
                return True
                
            except PermissionError as e:
                if attempt < max_retries - 1:
                    logger.warning(f'文件被占用，等待后重试 ({attempt + 1}/{max_retries}): {src}')
                    time.sleep(retry_delay)
                    retry_delay *= 2  # 指数退避
                else:
                    logger.error(f'移动文件失败，已达到最大重试次数: {src} -> {dst}')
                    raise e
            except Exception as e:
                logger.error(f'移动文件时发生错误: {src} -> {dst}, 错误: {e}')
                # 对于非权限错误，不重试，直接抛出
                raise e
        return False

    def make_report(self, app, devices, video, platform=Platform.Android, model='normal', cores=0):
        logger.info('Generating test results ...')
        current_time = time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())
        result_dict = {
            "app": app,
            "icon": "",
            "platform": platform,
            "model": model,
            "devices": devices,
            "ctime": current_time,
            "video": video,
            "cores":cores
        }
        content = json.dumps(result_dict)
        self.create_file(filename='result.json', content=content)
        report_new_dir = os.path.join(self.report_dir, f'apm_{current_time}')
        if not os.path.exists(report_new_dir):
            os.mkdir(report_new_dir)

        # 安全移动文件，特别处理可能被占用的录屏文件
        moved_files = []
        for f in os.listdir(self.report_dir):
            filename = os.path.join(self.report_dir, f)
            if f.split(".")[-1] in ['log', 'json', 'mkv']:
                # 检查文件是否真实存在且不是目录
                if not os.path.isfile(filename):
                    logger.warning(f'跳过非文件项: {filename}')
                    continue
                    
                try:
                    if f.endswith('.mkv'):
                        # 录屏文件可能被占用，使用安全移动方法
                        logger.info(f'移动录屏文件: {filename}')
                        self._safe_move_file(filename, report_new_dir)
                    else:
                        # 普通文件直接移动
                        dst_file = os.path.join(report_new_dir, f)
                        if os.path.exists(dst_file):
                            logger.warning(f'目标文件已存在，删除后重新移动: {dst_file}')
                            os.remove(dst_file)
                        shutil.move(filename, report_new_dir)
                    moved_files.append(f)
                except Exception as e:
                    logger.error(f'移动文件失败: {filename}, 错误: {e}')
                    # 继续处理其他文件，不中断整个流程
                    continue
        
        if moved_files:
            logger.info(f'成功移动文件: {moved_files}')
        else:
            logger.info('没有文件需要移动')
            
        logger.info('Generating test results success: {}'.format(report_new_dir))
        return f'apm_{current_time}'

    def instance_type(self, data):
        if isinstance(data, float):
            return 'float'
        elif isinstance(data, int):
            return 'int'
        else:
            return 'int'
    
    def open_file(self, path, mode):
        with open(path, mode) as f:
            for line in f:
                yield line
    
    def readJson(self, scene):
        path = os.path.join(self.report_dir,scene,'result.json')
        with open(file=path, mode='r', encoding='utf-8') as fp:
            result_dict = json.loads(fp.read())
        return result_dict

    def _setH5Perfs(self, scene):
        """聚合 H5 报告数据:运行时曲线(fps/heap/nodes/longtask) + 加载成绩单(取各日志末值)"""
        result_json = self.readJson(scene=scene)

        def series(fname):
            data, _, _ = self.readLog(scene=scene, filename=fname)
            return data  # [{x,y}]

        def last(fname):
            _, vals, _ = self.readLog(scene=scene, filename=fname)
            return vals[-1] if vals else 0

        def delta_series(fname):
            """把累计值日志(长任务/布局/样式重算)换算成'每次采样新增'(与实时页一致)"""
            out, prev = [], None
            for pt in series(fname):
                cur = pt['y']
                out.append({'x': pt['x'], 'y': 0 if (prev is None or cur < prev) else cur - prev})
                prev = cur
            return out

        # 热点剖析结果(若该会话开了实时剖析则有 h5_profile.json)
        profile_top = []
        try:
            ppath = os.path.join(self.report_dir, scene, 'h5_profile.json')
            if os.path.exists(ppath):
                with open(ppath, encoding='utf-8') as fp:
                    profile_top = json.loads(fp.read()).get('top', [])
        except Exception:
            pass
        # 最近长任务明细(归因表),来自 h5_longtasks.json
        longtasks = []
        try:
            lpath = os.path.join(self.report_dir, scene, 'h5_longtasks.json')
            if os.path.exists(lpath):
                with open(lpath, encoding='utf-8') as fp:
                    longtasks = json.loads(fp.read()).get('lt', [])
        except Exception:
            pass
        # 资源瀑布流,来自 h5_waterfall.json
        waterfall = []
        try:
            wpath = os.path.join(self.report_dir, scene, 'h5_waterfall.json')
            if os.path.exists(wpath):
                with open(wpath, encoding='utf-8') as fp:
                    waterfall = json.loads(fp.read()).get('resources', [])
        except Exception:
            pass

        return {
            'app': result_json.get('app'),
            'devices': result_json.get('devices'),
            'ctime': result_json.get('ctime'),
            'fps': series('h5_pagefps.log'),
            'heap': series('h5_heap.log'),
            'nodes': series('h5_nodes.log'),
            'longtask': delta_series('h5_longtask.log'),
            # 对标 Chrome Performance Monitor:监听器数(绝对) + 布局/样式重算(每次新增)
            'listeners': series('h5_listeners.log'),
            'layout': delta_series('h5_layout.log'),
            'recalc': delta_series('h5_recalc.log'),
            'inp': series('h5_inp.log'),
            'documents': series('h5_documents.log'),
            'frames': series('h5_frames.log'),
            'profile': profile_top,    # 热点函数表(可能为空)
            'longtasks': longtasks,    # 最近长任务归因(可能为空)
            'waterfall': waterfall,    # 资源瀑布流(可能为空)
            # 宿主进程原生指标(发热归因);H5-only 会话无这些日志时返回空列表
            'hostCpuApp': series('h5_host_cpu_app.log'),
            'hostCpuSys': series('h5_host_cpu_sys.log'),
            'hostGpu': series('h5_host_gpu.log'),
            'hostTemp': series('h5_host_temp.log'),
            'load': {
                'fp': last('h5_fp.log'), 'lcp': last('h5_lcp.log'), 'fcp': last('h5_fcp.log'),
                'ttfb': last('h5_ttfb.log'), 'dcl': last('h5_dcl.log'), 'load': last('h5_load.log'),
                'tbt': last('h5_tbt.log'), 'tti': last('h5_tti.log'),
            },
        }

    def readLog(self, scene, filename, max_points=0):
        """
        Read apmlog file data with optional downsampling

        Args:
            scene: 场景名称
            filename: 日志文件名
            max_points: 最大数据点数，0 表示不采样，默认 0

        Returns:
            (log_data_list, target_data_list, total_points) 三元组
            - log_data_list: [{"x": timestamp, "y": value}, ...]
            - target_data_list: [value, ...]
            - total_points: 原始数据点总数
        """
        full_path = os.path.join(self.report_dir, scene, filename)
        if not os.path.exists(full_path):
            return list(), list(), 0

        # mtime-aware 缓存：文件未变更时复用全量解析结果，避免同一文件被重复读盘与解析
        mtime = os.path.getmtime(full_path)
        cache_key = (scene, filename)
        cached = self._log_cache.get(cache_key)
        if cached is not None and cached[0] == mtime:
            log_data_list, target_data_list = cached[1], cached[2]
            self._log_cache.move_to_end(cache_key)
        else:
            log_data_list, target_data_list = self._parse_log_file(full_path)
            self._log_cache[cache_key] = (mtime, log_data_list, target_data_list)
            self._log_cache.move_to_end(cache_key)
            while len(self._log_cache) > self._log_cache_limit:
                self._log_cache.popitem(last=False)

        # 记录原始数据点数量
        total_points = len(log_data_list)

        # 应用 LTTB 降采样（返回新列表，不污染缓存中的全量数据）
        if max_points > 0 and total_points > max_points:
            log_data_list = downsample_lttb(log_data_list, max_points)
            # 重建 target_data_list
            target_data_list = [item['y'] for item in log_data_list]

        return log_data_list, target_data_list, total_points

    def _parse_log_file(self, full_path):
        """逐行解析单个 log 文件，返回 (log_data_list, target_data_list) 全量数据"""
        log_data_list = list()
        target_data_list = list()
        for line in self.open_file(full_path, "r"):
            if '=' not in line:
                continue
            parts = line.split('=')
            raw = parts[1].strip()
            if not raw:
                continue
            try:
                value = int(raw) if raw.lstrip('-').isdigit() else float(raw)
            except ValueError:
                continue
            log_data_list.append({"x": parts[0].strip(), "y": value})
            target_data_list.append(value)
        return log_data_list, target_data_list
        
    def getCpuLog(self, platform, scene, max_points=0):
        targetDic = dict()
        cpu_app_data, _, cpu_app_total = self.readLog(scene=scene, filename='cpu_app.log', max_points=0)
        cpu_sys_data, _, cpu_sys_total = self.readLog(scene=scene, filename='cpu_sys.log', max_points=0)
        if max_points > 0:
            cpu_app_data, cpu_sys_data = downsample_lttb_multi([cpu_app_data, cpu_sys_data], max_points)
        targetDic['cpuAppData'] = cpu_app_data
        targetDic['cpuSysData'] = cpu_sys_data
        result = {
            'status': 1,
            'cpuAppData': targetDic['cpuAppData'],
            'cpuSysData': targetDic['cpuSysData'],
            'meta': {
                'sampled': max_points > 0 and max(cpu_app_total, cpu_sys_total) > max_points,
                'max_points': max_points,
                'total_points': max(cpu_app_total, cpu_sys_total)
            }
        }
        return result
    
    def getCpuLogCompare(self, platform, scene1, scene2, max_points=0):
        targetDic = dict()
        scene1_data, _, scene1_total = self.readLog(scene=scene1, filename='cpu_app.log', max_points=max_points)
        scene2_data, _, scene2_total = self.readLog(scene=scene2, filename='cpu_app.log', max_points=max_points)
        targetDic['scene1'] = scene1_data
        targetDic['scene2'] = scene2_data
        result = {
            'status': 1,
            'scene1': targetDic['scene1'],
            'scene2': targetDic['scene2'],
            'meta': {
                'sampled': max_points > 0 and max(scene1_total, scene2_total) > max_points,
                'max_points': max_points,
                'total_points': max(scene1_total, scene2_total)
            }
        }
        return result
    
    def getGpuLog(self, platform, scene, max_points=0):
        targetDic = dict()
        gpu_data, _, total_points = self.readLog(scene=scene, filename='gpu.log', max_points=max_points)
        targetDic['gpu'] = gpu_data
        result = {
            'status': 1,
            'gpu': targetDic['gpu'],
            'meta': {
                'sampled': max_points > 0 and total_points > max_points,
                'max_points': max_points,
                'total_points': total_points
            }
        }
        return result
    
    def getGpuLogCompare(self, platform, scene1, scene2, max_points=0):
        targetDic = dict()
        scene1_data, _, scene1_total = self.readLog(scene=scene1, filename='gpu.log', max_points=max_points)
        scene2_data, _, scene2_total = self.readLog(scene=scene2, filename='gpu.log', max_points=max_points)
        targetDic['scene1'] = scene1_data
        targetDic['scene2'] = scene2_data
        result = {
            'status': 1,
            'scene1': targetDic['scene1'],
            'scene2': targetDic['scene2'],
            'meta': {
                'sampled': max_points > 0 and max(scene1_total, scene2_total) > max_points,
                'max_points': max_points,
                'total_points': max(scene1_total, scene2_total)
            }
        }
        return result
    
    def getMemLog(self, platform, scene, max_points=0):
        targetDic = dict()
        if platform == Platform.Android:
            mem_total_data, _, mem_total_total = self.readLog(scene=scene, filename='mem_total.log', max_points=0)
            mem_swap_data, _, mem_swap_total = self.readLog(scene=scene, filename='mem_swap.log', max_points=0)
            if max_points > 0:
                mem_total_data, mem_swap_data = downsample_lttb_multi([mem_total_data, mem_swap_data], max_points)
            total_points = max(mem_total_total, mem_swap_total)
            result = {
                'status': 1,
                'memTotalData': mem_total_data,
                'memSwapData': mem_swap_data,
                'meta': {
                    'sampled': max_points > 0 and total_points > max_points,
                    'max_points': max_points,
                    'total_points': total_points
                }
            }
        else:
            mem_total_data, _, mem_total_total = self.readLog(scene=scene, filename='mem_total.log', max_points=max_points)
            total_points = mem_total_total
            result = {
                'status': 1,
                'memTotalData': mem_total_data,
                'meta': {
                    'sampled': max_points > 0 and total_points > max_points,
                    'max_points': max_points,
                    'total_points': total_points
                }
            }
        return result
    
    def getMemDetailLog(self, platform, scene, max_points=0):
        targetDic = dict()
        total_points_list = []
        keys, series = [], []
        for key, filename in [
            ('java_heap', 'mem_java_heap.log'),
            ('native_heap', 'mem_native_heap.log'),
            ('code_pss', 'mem_code_pss.log'),
            ('stack_pss', 'mem_stack_pss.log'),
            ('graphics_pss', 'mem_graphics_pss.log'),
            ('private_pss', 'mem_private_pss.log'),
            ('system_pss', 'mem_system_pss.log')
        ]:
            data, _, total = self.readLog(scene=scene, filename=filename, max_points=0)
            keys.append(key)
            series.append(data)
            total_points_list.append(total)
        if max_points > 0:
            series = downsample_lttb_multi(series, max_points)
        for key, data in zip(keys, series):
            targetDic[key] = data
        max_total = max(total_points_list) if total_points_list else 0
        result = {
            'status': 1,
            'memory_detail': targetDic,
            'meta': {
                'sampled': max_points > 0 and max_total > max_points,
                'max_points': max_points,
                'total_points': max_total
            }
        }
        return result
    
    def getCpuCoreLog(self, platform, scene, max_points=0):
        targetDic = dict()
        cores = self.readJson(scene=scene).get('cores', 0)
        total_points_list = []
        if int(cores) > 0:
            keys, series = [], []
            for i in range(int(cores)):
                data, _, total = self.readLog(scene=scene, filename='cpu{}.log'.format(i), max_points=0)
                keys.append('cpu{}'.format(i))
                series.append(data)
                total_points_list.append(total)
            if max_points > 0:
                series = downsample_lttb_multi(series, max_points)
            for key, data in zip(keys, series):
                targetDic[key] = data
        max_total = max(total_points_list) if total_points_list else 0
        result = {
            'status': 1,
            'cores': cores,
            'cpu_core': targetDic,
            'meta': {
                'sampled': max_points > 0 and max_total > max_points,
                'max_points': max_points,
                'total_points': max_total
            }
        }
        return result
    
    def getMemLogCompare(self, platform, scene1, scene2, max_points=0):
        targetDic = dict()
        scene1_data, _, scene1_total = self.readLog(scene=scene1, filename='mem_total.log', max_points=max_points)
        scene2_data, _, scene2_total = self.readLog(scene=scene2, filename='mem_total.log', max_points=max_points)
        targetDic['scene1'] = scene1_data
        targetDic['scene2'] = scene2_data
        result = {
            'status': 1,
            'scene1': targetDic['scene1'],
            'scene2': targetDic['scene2'],
            'meta': {
                'sampled': max_points > 0 and max(scene1_total, scene2_total) > max_points,
                'max_points': max_points,
                'total_points': max(scene1_total, scene2_total)
            }
        }
        return result
    
    def getBatteryLog(self, platform, scene, max_points=0):
        targetDic = dict()
        total_points_list = []
        if platform == Platform.Android:
            level_data, _, level_total = self.readLog(scene=scene, filename='battery_level.log', max_points=0)
            tem_data, _, tem_total = self.readLog(scene=scene, filename='battery_tem.log', max_points=0)
            if max_points > 0:
                level_data, tem_data = downsample_lttb_multi([level_data, tem_data], max_points)
            targetDic['batteryLevel'] = level_data
            targetDic['batteryTem'] = tem_data
            total_points_list = [level_total, tem_total]
            result = {
                'status': 1,
                'batteryLevel': targetDic['batteryLevel'],
                'batteryTem': targetDic['batteryTem'],
                'meta': {
                    'sampled': max_points > 0 and max(total_points_list) > max_points,
                    'max_points': max_points,
                    'total_points': max(total_points_list)
                }
            }
        else:
            tem_data, _, tem_total = self.readLog(scene=scene, filename='battery_tem.log', max_points=0)
            current_data, _, current_total = self.readLog(scene=scene, filename='battery_current.log', max_points=0)
            voltage_data, _, voltage_total = self.readLog(scene=scene, filename='battery_voltage.log', max_points=0)
            power_data, _, power_total = self.readLog(scene=scene, filename='battery_power.log', max_points=0)
            if max_points > 0:
                tem_data, current_data, voltage_data, power_data = downsample_lttb_multi(
                    [tem_data, current_data, voltage_data, power_data], max_points)
            targetDic['batteryTem'] = tem_data
            targetDic['batteryCurrent'] = current_data
            targetDic['batteryVoltage'] = voltage_data
            targetDic['batteryPower'] = power_data
            total_points_list = [tem_total, current_total, voltage_total, power_total]
            result = {
                'status': 1,
                'batteryTem': targetDic['batteryTem'],
                'batteryCurrent': targetDic['batteryCurrent'],
                'batteryVoltage': targetDic['batteryVoltage'],
                'batteryPower': targetDic['batteryPower'],
                'meta': {
                    'sampled': max_points > 0 and max(total_points_list) > max_points,
                    'max_points': max_points,
                    'total_points': max(total_points_list)
                }
            }
        return result
    
    def getBatteryLogCompare(self, platform, scene1, scene2, max_points=0):
        targetDic = dict()
        if platform == Platform.Android:
            scene1_data, _, scene1_total = self.readLog(scene=scene1, filename='battery_level.log', max_points=max_points)
            scene2_data, _, scene2_total = self.readLog(scene=scene2, filename='battery_level.log', max_points=max_points)
        else:
            scene1_data, _, scene1_total = self.readLog(scene=scene1, filename='battery_power.log', max_points=max_points)
            scene2_data, _, scene2_total = self.readLog(scene=scene2, filename='battery_power.log', max_points=max_points)
        targetDic['scene1'] = scene1_data
        targetDic['scene2'] = scene2_data
        result = {
            'status': 1,
            'scene1': targetDic['scene1'],
            'scene2': targetDic['scene2'],
            'meta': {
                'sampled': max_points > 0 and max(scene1_total, scene2_total) > max_points,
                'max_points': max_points,
                'total_points': max(scene1_total, scene2_total)
            }
        }
        return result
    
    def getFlowLog(self, platform, scene, max_points=0):
        targetDic = dict()
        up_data, _, up_total = self.readLog(scene=scene, filename='upflow.log', max_points=0)
        down_data, _, down_total = self.readLog(scene=scene, filename='downflow.log', max_points=0)
        if max_points > 0:
            up_data, down_data = downsample_lttb_multi([up_data, down_data], max_points)
        targetDic['upFlow'] = up_data
        targetDic['downFlow'] = down_data
        max_total = max(up_total, down_total)
        result = {
            'status': 1,
            'upFlow': targetDic['upFlow'],
            'downFlow': targetDic['downFlow'],
            'meta': {
                'sampled': max_points > 0 and max_total > max_points,
                'max_points': max_points,
                'total_points': max_total
            }
        }
        return result
    
    def getFlowSendLogCompare(self, platform, scene1, scene2, max_points=0):
        targetDic = dict()
        scene1_data, _, scene1_total = self.readLog(scene=scene1, filename='upflow.log', max_points=max_points)
        scene2_data, _, scene2_total = self.readLog(scene=scene2, filename='upflow.log', max_points=max_points)
        targetDic['scene1'] = scene1_data
        targetDic['scene2'] = scene2_data
        result = {
            'status': 1,
            'scene1': targetDic['scene1'],
            'scene2': targetDic['scene2'],
            'meta': {
                'sampled': max_points > 0 and max(scene1_total, scene2_total) > max_points,
                'max_points': max_points,
                'total_points': max(scene1_total, scene2_total)
            }
        }
        return result
    
    def getFlowRecvLogCompare(self, platform, scene1, scene2, max_points=0):
        targetDic = dict()
        scene1_data, _, scene1_total = self.readLog(scene=scene1, filename='downflow.log', max_points=max_points)
        scene2_data, _, scene2_total = self.readLog(scene=scene2, filename='downflow.log', max_points=max_points)
        targetDic['scene1'] = scene1_data
        targetDic['scene2'] = scene2_data
        result = {
            'status': 1,
            'scene1': targetDic['scene1'],
            'scene2': targetDic['scene2'],
            'meta': {
                'sampled': max_points > 0 and max(scene1_total, scene2_total) > max_points,
                'max_points': max_points,
                'total_points': max(scene1_total, scene2_total)
            }
        }
        return result
    
    def getFpsLog(self, platform, scene, max_points=0):
        targetDic = dict()
        if platform == Platform.Android:
            fps_data, _, fps_total = self.readLog(scene=scene, filename='fps.log', max_points=0)
            jank_data, _, jank_total = self.readLog(scene=scene, filename='jank.log', max_points=0)
            if max_points > 0:
                fps_data, jank_data = downsample_lttb_multi([fps_data, jank_data], max_points)
            total_points = max(fps_total, jank_total)
            result = {
                'status': 1,
                'fps': fps_data,
                'jank': jank_data,
                'meta': {
                    'sampled': max_points > 0 and total_points > max_points,
                    'max_points': max_points,
                    'total_points': total_points
                }
            }
        else:
            fps_data, _, fps_total = self.readLog(scene=scene, filename='fps.log', max_points=max_points)
            total_points = fps_total
            result = {
                'status': 1,
                'fps': fps_data,
                'meta': {
                    'sampled': max_points > 0 and total_points > max_points,
                    'max_points': max_points,
                    'total_points': total_points
                }
            }
        return result
    
    def getDiskLog(self, platform, scene, max_points=0):
        targetDic = dict()
        used_data, _, used_total = self.readLog(scene=scene, filename='disk_used.log', max_points=0)
        free_data, _, free_total = self.readLog(scene=scene, filename='disk_free.log', max_points=0)
        if max_points > 0:
            used_data, free_data = downsample_lttb_multi([used_data, free_data], max_points)
        targetDic['used'] = used_data
        targetDic['free'] = free_data
        max_total = max(used_total, free_total)
        result = {
            'status': 1,
            'used': targetDic['used'],
            'free': targetDic['free'],
            'meta': {
                'sampled': max_points > 0 and max_total > max_points,
                'max_points': max_points,
                'total_points': max_total
            }
        }
        return result

    def analysisDisk(self, scene):
        initail_disk_list = list()
        current_disk_list = list()
        sum_init_disk = dict()
        sum_current_disk = dict()
        if os.path.exists(os.path.join(self.report_dir,scene,'initail_disk.log')):
            size_list = list()
            used_list = list()
            free_list = list()
            lines = self.open_file(os.path.join(self.report_dir,scene,'initail_disk.log'), "r")
            for line in lines:
                if 'Filesystem' not in line and line.strip() != '':
                    disk_value_list = line.split()
                    disk_dict = dict(
                        filesystem = disk_value_list[0],
                        blocks = disk_value_list[1],
                        used = disk_value_list[2],
                        available = disk_value_list[3],
                        use_percent = disk_value_list[4],
                        mounted = disk_value_list[5]
                    )
                    initail_disk_list.append(disk_dict)
                    size_list.append(int(disk_value_list[1]))
                    used_list.append(int(disk_value_list[2]))
                    free_list.append(int(disk_value_list[3]))
            sum_init_disk['sum_size'] = int(sum(size_list) / 1024 / 1024)
            sum_init_disk['sum_used'] = int(sum(used_list) / 1024)
            sum_init_disk['sum_free'] = int(sum(free_list) / 1024)
               
        if os.path.exists(os.path.join(self.report_dir,scene,'current_disk.log')):
            size_list = list()
            used_list = list()
            free_list = list()
            lines = self.open_file(os.path.join(self.report_dir,scene,'current_disk.log'), "r")
            for line in lines:
                if 'Filesystem' not in line and line.strip() != '':
                    disk_value_list = line.split()
                    disk_dict = dict(
                        filesystem = disk_value_list[0],
                        blocks = disk_value_list[1],
                        used = disk_value_list[2],
                        available = disk_value_list[3],
                        use_percent = disk_value_list[4],
                        mounted = disk_value_list[5]
                    )
                    current_disk_list.append(disk_dict)
                    size_list.append(int(disk_value_list[1]))
                    used_list.append(int(disk_value_list[2]))
                    free_list.append(int(disk_value_list[3]))
            sum_current_disk['sum_size'] = int(sum(size_list) / 1024 / 1024)
            sum_current_disk['sum_used'] = int(sum(used_list) / 1024)
            sum_current_disk['sum_free'] = int(sum(free_list) / 1024)       
                 
        return initail_disk_list, current_disk_list, sum_init_disk, sum_current_disk

    def getFpsLogCompare(self, platform, scene1, scene2, max_points=0):
        targetDic = dict()
        scene1_data, _, scene1_total = self.readLog(scene=scene1, filename='fps.log', max_points=max_points)
        scene2_data, _, scene2_total = self.readLog(scene=scene2, filename='fps.log', max_points=max_points)
        targetDic['scene1'] = scene1_data
        targetDic['scene2'] = scene2_data
        result = {
            'status': 1,
            'scene1': targetDic['scene1'],
            'scene2': targetDic['scene2'],
            'meta': {
                'sampled': max_points > 0 and max(scene1_total, scene2_total) > max_points,
                'max_points': max_points,
                'total_points': max(scene1_total, scene2_total)
            }
        }
        return result
        
    def approximateSize(self, size, a_kilobyte_is_1024_bytes=True):
        '''
        convert a file size to human-readable form.
        Keyword arguments:
        size -- file size in bytes
        a_kilobyte_is_1024_bytes -- if True (default),use multiples of 1024
                                    if False, use multiples of 1000
        Returns: string
        '''

        suffixes = {1000: ['KB', 'MB', 'GB', 'TB', 'PB', 'EB', 'ZB', 'YB'],
                    1024: ['KiB', 'MiB', 'GiB', 'TiB', 'PiB', 'EiB', 'ZiB', 'YiB']}

        if size < 0:
            raise ValueError('number must be non-negative')

        multiple = 1024 if a_kilobyte_is_1024_bytes else 1000

        suffix = suffixes[multiple][0]
        for suffix in suffixes[multiple]:
            size /= multiple
            if size < multiple:
                return '{0:.2f} {1}'.format(size, suffix)
        # 超过最大后缀时用最大单位兜底,避免返回 None
        return '{0:.2f} {1}'.format(size, suffix)

    def _setAndroidPerfs(self, scene):
        """Aggregate APM data for Android"""

        result_json = self.readJson(scene=scene)
        app = result_json.get('app')
        devices = result_json.get('devices')
        platform = result_json.get('platform')
        ctime = result_json.get('ctime')

        _, cpuAppData, _ = self.readLog(scene=scene, filename=f'cpu_app.log')
        _, cpuSystemData, _ = self.readLog(scene=scene, filename=f'cpu_sys.log')
        if cpuAppData.__len__() > 0 and cpuSystemData.__len__() > 0:
            cpuAppRate = f'{round(sum(cpuAppData) / len(cpuAppData), 2)}%'
            cpuSystemRate = f'{round(sum(cpuSystemData) / len(cpuSystemData), 2)}%'
        else:
            cpuAppRate, cpuSystemRate = 0, 0

        _, batteryLevelData, _ = self.readLog(scene=scene, filename=f'battery_level.log')
        _, batteryTemlData, _ = self.readLog(scene=scene, filename=f'battery_tem.log')
        if batteryLevelData.__len__() > 0 and batteryTemlData.__len__() > 0:
            batteryLevel = f'{batteryLevelData[-1]}%'
            batteryTeml = f'{batteryTemlData[-1]}°C'
        else:
            batteryLevel, batteryTeml = 0, 0


        _, totalPassData, _ = self.readLog(scene=scene, filename=f'mem_total.log')

        if totalPassData.__len__() > 0:
            _, swapPassData, _ = self.readLog(scene=scene, filename=f'mem_swap.log')
            totalPassAvg = f'{round(sum(totalPassData) / len(totalPassData), 2)}MB'
            swapPassAvg = f'{round(sum(swapPassData) / len(swapPassData), 2)}MB' if swapPassData else '0MB'
        else:
            totalPassAvg, swapPassAvg = 0, 0

        _, fpsData, _ = self.readLog(scene=scene, filename=f'fps.log')
        _, jankData, _ = self.readLog(scene=scene, filename=f'jank.log')
        if fpsData.__len__() > 0:
            fpsAvg = f'{int(sum(fpsData) / len(fpsData))}HZ/s'
            jankAvg = f'{int(sum(jankData))}'
        else:
            fpsAvg, jankAvg = 0, 0

        pre_net_path = os.path.join(self.report_dir, scene, 'pre_net.json')
        end_net_path = os.path.join(self.report_dir, scene, 'end_net.json')
        if os.path.exists(pre_net_path) and os.path.exists(end_net_path):
            with open(pre_net_path) as f_pre, open(end_net_path) as f_end:
                json_pre = json.loads(f_pre.read())
                json_end = json.loads(f_end.read())
                send = json_end['send'] - json_pre['send']
                recv = json_end['recv'] - json_pre['recv']
        else:
            send, recv = 0, 0
        flowSend = f'{round(float(send / 1024), 2)}MB'
        flowRecv = f'{round(float(recv / 1024), 2)}MB'

        _, gpuData, _ = self.readLog(scene=scene, filename='gpu.log')
        if gpuData.__len__() > 0:
            gpu = round(sum(gpuData) / len(gpuData), 2)
        else:
            gpu = 0

        mem_detail_flag = os.path.exists(os.path.join(self.report_dir,scene,'mem_java_heap.log'))
        disk_flag = os.path.exists(os.path.join(self.report_dir,scene,'disk_free.log'))
        thermal_flag = os.path.exists(os.path.join(self.report_dir,scene,'init_thermal_temp.json'))
        cpu_core_flag = os.path.exists(os.path.join(self.report_dir,scene,'cpu0.log'))
        apm_dict = dict()
        apm_dict['app'] = app
        apm_dict['devices'] = devices
        apm_dict['platform'] = platform
        apm_dict['ctime'] = ctime
        apm_dict['cpuAppRate'] = cpuAppRate
        apm_dict['cpuSystemRate'] = cpuSystemRate
        apm_dict['totalPassAvg'] = totalPassAvg
        apm_dict['swapPassAvg'] = swapPassAvg
        apm_dict['fps'] = fpsAvg
        apm_dict['jank'] = jankAvg
        apm_dict['flow_send'] = flowSend
        apm_dict['flow_recv'] = flowRecv
        apm_dict['batteryLevel'] = batteryLevel
        apm_dict['batteryTeml'] = batteryTeml
        apm_dict['mem_detail_flag'] = mem_detail_flag
        apm_dict['disk_flag'] = disk_flag
        apm_dict['gpu'] = gpu
        apm_dict['thermal_flag'] = thermal_flag
        apm_dict['cpu_core_flag'] = cpu_core_flag
        
        if thermal_flag:
            init_path = os.path.join(self.report_dir, scene, 'init_thermal_temp.json')
            current_path = os.path.join(self.report_dir, scene, 'current_thermal_temp.json')
            with open(init_path, encoding='utf-8') as _f:
                init_thermal_temp = json.loads(_f.read())
            if os.path.exists(current_path):
                with open(current_path, encoding='utf-8') as _f:
                    current_thermal_temp = json.loads(_f.read())
            else:
                current_thermal_temp = init_thermal_temp
            apm_dict['init_thermal_temp'] = init_thermal_temp
            apm_dict['current_thermal_temp'] = current_thermal_temp

        return apm_dict

    def _setiOSPerfs(self, scene):
        """Aggregate APM data for iOS"""

        result_json = self.readJson(scene=scene)
        app = result_json.get('app')
        devices = result_json.get('devices')
        platform = result_json.get('platform')
        ctime = result_json.get('ctime')

        _, cpuAppData, _ = self.readLog(scene=scene, filename=f'cpu_app.log')
        _, cpuSystemData, _ = self.readLog(scene=scene, filename=f'cpu_sys.log')
        if cpuAppData.__len__() > 0 and cpuSystemData.__len__() > 0:
            cpuAppRate = f'{round(sum(cpuAppData) / len(cpuAppData), 2)}%'
            cpuSystemRate = f'{round(sum(cpuSystemData) / len(cpuSystemData), 2)}%'
        else:
            cpuAppRate, cpuSystemRate = 0, 0

        _, totalPassData, _ = self.readLog(scene=scene, filename='mem_total.log')
        if totalPassData.__len__() > 0:
            totalPassAvg = f'{round(sum(totalPassData) / len(totalPassData), 2)}MB'
        else:
            totalPassAvg = 0

        _, fpsData, _ = self.readLog(scene=scene, filename='fps.log')
        if fpsData.__len__() > 0:
            fpsAvg = f'{int(sum(fpsData) / len(fpsData))}HZ/s'
        else:
            fpsAvg = 0

        _, flowSendData, _ = self.readLog(scene=scene, filename='upflow.log')
        _, flowRecvData, _ = self.readLog(scene=scene, filename='downflow.log')
        if flowSendData.__len__() > 0:
            flowSend = f'{round(float(sum(flowSendData) / 1024), 2)}MB'
            flowRecv = f'{round(float(sum(flowRecvData) / 1024), 2)}MB'
        else:
            flowSend, flowRecv = 0, 0

        _, batteryTemlData, _ = self.readLog(scene=scene, filename='battery_tem.log')
        _, batteryCurrentData, _ = self.readLog(scene=scene, filename='battery_current.log')
        _, batteryVoltageData, _ = self.readLog(scene=scene, filename='battery_voltage.log')
        _, batteryPowerData, _ = self.readLog(scene=scene, filename='battery_power.log')
        if batteryTemlData.__len__() > 0:
            batteryTeml = int(batteryTemlData[-1])
            batteryCurrent = int(sum(batteryCurrentData) / len(batteryCurrentData)) if batteryCurrentData else 0
            batteryVoltage = int(sum(batteryVoltageData) / len(batteryVoltageData)) if batteryVoltageData else 0
            batteryPower = int(sum(batteryPowerData) / len(batteryPowerData)) if batteryPowerData else 0
        else:
            batteryTeml, batteryCurrent, batteryVoltage, batteryPower = 0, 0, 0, 0

        _, gpuData, _ = self.readLog(scene=scene, filename='gpu.log')
        if gpuData.__len__() > 0:
            gpu = round(sum(gpuData) / len(gpuData), 2)
        else:
            gpu = 0
        disk_flag = os.path.exists(os.path.join(self.report_dir, scene, 'disk_free.log'))
        apm_dict = dict()
        apm_dict['app'] = app
        apm_dict['devices'] = devices
        apm_dict['platform'] = platform
        apm_dict['ctime'] = ctime
        apm_dict['cpuAppRate'] = cpuAppRate
        apm_dict['cpuSystemRate'] = cpuSystemRate
        apm_dict['totalPassAvg'] = totalPassAvg
        apm_dict['nativePassAvg'] = 0
        apm_dict['dalvikPassAvg'] = 0
        apm_dict['fps'] = fpsAvg
        apm_dict['jank'] = 0
        apm_dict['flow_send'] = flowSend
        apm_dict['flow_recv'] = flowRecv
        apm_dict['batteryTeml'] = batteryTeml
        apm_dict['batteryCurrent'] = batteryCurrent
        apm_dict['batteryVoltage'] = batteryVoltage
        apm_dict['batteryPower'] = batteryPower
        apm_dict['gpu'] = gpu
        apm_dict['disk_flag'] = disk_flag
        return apm_dict

    def _setpkPerfs(self, scene):
        """Aggregate APM data for pk model"""
        _, cpuAppData1, _ = self.readLog(scene=scene, filename='cpu_app1.log')
        cpuAppRate1 = f'{round(sum(cpuAppData1) / len(cpuAppData1), 2)}%' if cpuAppData1 else '0%'
        _, cpuAppData2, _ = self.readLog(scene=scene, filename='cpu_app2.log')
        cpuAppRate2 = f'{round(sum(cpuAppData2) / len(cpuAppData2), 2)}%' if cpuAppData2 else '0%'

        _, totalPassData1, _ = self.readLog(scene=scene, filename='mem1.log')
        totalPassAvg1 = f'{round(sum(totalPassData1) / len(totalPassData1), 2)}MB' if totalPassData1 else '0MB'
        _, totalPassData2, _ = self.readLog(scene=scene, filename='mem2.log')
        totalPassAvg2 = f'{round(sum(totalPassData2) / len(totalPassData2), 2)}MB' if totalPassData2 else '0MB'

        _, fpsData1, _ = self.readLog(scene=scene, filename='fps1.log')
        fpsAvg1 = f'{int(sum(fpsData1) / len(fpsData1))}HZ/s' if fpsData1 else '0HZ/s'
        _, fpsData2, _ = self.readLog(scene=scene, filename='fps2.log')
        fpsAvg2 = f'{int(sum(fpsData2) / len(fpsData2))}HZ/s' if fpsData2 else '0HZ/s'

        _, networkData1, _ = self.readLog(scene=scene, filename='network1.log')
        network1 = f'{round(float(sum(networkData1) / 1024), 2)}MB'
        _, networkData2, _ = self.readLog(scene=scene, filename='network2.log')
        network2 = f'{round(float(sum(networkData2) / 1024), 2)}MB'

        apm_dict = dict()
        apm_dict['cpuAppRate1'] = cpuAppRate1
        apm_dict['cpuAppRate2'] = cpuAppRate2
        apm_dict['totalPassAvg1'] = totalPassAvg1
        apm_dict['totalPassAvg2'] = totalPassAvg2
        apm_dict['network1'] = network1
        apm_dict['network2'] = network2
        apm_dict['fpsAvg1'] = fpsAvg1
        apm_dict['fpsAvg2'] = fpsAvg2
        return apm_dict

class Method:
    
    @classmethod
    def _request(cls, request, object):
        match(request.method):
            case 'POST':
                return request.form[object]
            case 'GET':
                return request.args[object]
            case _:
                raise Exception('request method error')
    
    @classmethod   
    def _setValue(cls, value, default = 0):
        # 实参在调用前已求值,原 try 无法捕获计算异常;此处仅对 None/无效值做兜底
        if value is None:
            return default
        return value
    
    @classmethod
    def _settings(cls, request):
        content = {}
        content['cpuWarning'] = (0, request.cookies.get('cpuWarning'))[request.cookies.get('cpuWarning') not in [None, 'NaN']]
        content['memWarning'] = (0, request.cookies.get('memWarning'))[request.cookies.get('memWarning') not in [None, 'NaN']]
        content['fpsWarning'] = (0, request.cookies.get('fpsWarning'))[request.cookies.get('fpsWarning') not in [None, 'NaN']]
        content['netdataRecvWarning'] = (0, request.cookies.get('netdataRecvWarning'))[request.cookies.get('netdataRecvWarning') not in [None, 'NaN']]
        content['netdataSendWarning'] = (0, request.cookies.get('netdataSendWarning'))[request.cookies.get('netdataSendWarning') not in [None, 'NaN']]
        content['betteryWarning'] = (0, request.cookies.get('betteryWarning'))[request.cookies.get('betteryWarning') not in [None, 'NaN']]
        content['gpuWarning'] = (0, request.cookies.get('gpuWarning'))[request.cookies.get('gpuWarning') not in [None, 'NaN']]
        content['duration'] = (0, request.cookies.get('duration'))[request.cookies.get('duration') not in [None, 'NaN']]
        content['magnax_host'] = ('', request.cookies.get('magnax_host'))[request.cookies.get('magnax_host') not in [None, 'NaN']]
        content['host_switch'] = request.cookies.get('host_switch')
        return content
    
    @classmethod
    def _index(cls, target: list, index: int, default: any):
        try:
            return target[index]
        except IndexError:
            return default

class Install:

    def uploadFile(self, file_path, file_obj):
        """save upload file"""
        try:
            file_obj.save(file_path)
            return True
        except Exception as e:
            logger.exception(e)
            return False            

    def downloadLink(self,filelink=None, path=None, name=None):
        try:
            logger.info('Install link : {}'.format(filelink))
            ssl._create_default_https_context = ssl._create_unverified_context
            file_size = int(urlopen(filelink).info().get('Content-Length', -1))
            header = {"Range": "bytes=%s-%s" % (0, file_size)}
            pbar = tqdm(
                total=file_size, initial=0,
                unit='B', unit_scale=True, desc=filelink.split('/')[-1])
            req = requests.get(filelink, headers=header, stream=True)
            with(open(os.path.join(path, name), 'wb')) as f:
                for chunk in req.iter_content(chunk_size=1024):
                    if chunk:
                         f.write(chunk)
                         pbar.update(len(chunk))
            pbar.close()
            return True
        except Exception as e:
            logger.exception(e)
            return False

    def installAPK(self, path):
        result = adb.shell_noDevice(cmd='install -r {}'.format(path))
        if result == 0:
            os.remove(path)
            return True, result
        else:
            return False, result

    def installIPA(self, path, device_id=None):
        """使用 pymobiledevice3 安装 IPA"""
        try:
            if not PMD3_AVAILABLE:
                logger.error("pymobiledevice3 not available, cannot install IPA")
                return False, -1

            from pymobiledevice3.services.installation_proxy import InstallationProxyService

            # 获取设备列表，如果没有指定设备则使用第一个
            if device_id is None:
                devices = _pmd3_list_devices_sync()
                if not devices:
                    logger.error("No iOS device connected")
                    return False, -1
                device_id = devices[0].serial

            lockdown_client = create_using_usbmux(serial=device_id)
            if lockdown_client is None:
                logger.error("Failed to connect to device")
                return False, -1

            # 使用 InstallationProxyService 安装 IPA
            installation = InstallationProxyService(lockdown=lockdown_client)
            installation.install_from_local(path)

            logger.info(f"Successfully installed IPA: {path}")
            os.remove(path)
            return True, 0

        except Exception as e:
            logger.error(f"Failed to install IPA: {e}")
            return False, -1

class Scrcpy:

    STATICPATH = os.path.dirname(os.path.realpath(__file__))
    DEFAULT_SCRCPY_PATH = {
        "64": os.path.join(STATICPATH, "scrcpy", "scrcpy-win64-v2.4", "scrcpy.exe"),
        "32": os.path.join(STATICPATH, "scrcpy", "scrcpy-win32-v2.4", "scrcpy.exe"),
        "default":"scrcpy"
    }
    
    @classmethod
    def scrcpy_path(cls):
        bit = platform.architecture()[0]
        path = cls.DEFAULT_SCRCPY_PATH["default"]
        if platform.system().lower().__contains__('windows'):
            if bit.__contains__('64'):
                path =  cls.DEFAULT_SCRCPY_PATH["64"]
            elif bit.__contains__('32'):
                path =  cls.DEFAULT_SCRCPY_PATH["32"]
        return path
    
    @classmethod
    def start_record(cls, device):
        f = File()
        logger.info('start record screen')
        win_cmd = "start /b {scrcpy_path} -s {deviceId} --no-playback --no-power-on --record={video}".format(
            scrcpy_path = cls.scrcpy_path(), 
            deviceId = device, 
            video = os.path.join(f.report_dir, 'record.mkv')
        )
        mac_cmd = "nohup {scrcpy_path} -s {deviceId} --no-playback --no-power-on --record={video} &".format(
            scrcpy_path = cls.scrcpy_path(), 
            deviceId = device, 
            video = os.path.join(f.report_dir, 'record.mkv')
        )
        if platform.system().lower().__contains__('windows'):
            result = os.system(win_cmd)
        else:
            result = os.system(mac_cmd)    
        if result == 0:
            logger.info("record screen success : {}".format(os.path.join(f.report_dir, 'record.mkv')))
        else:
            logger.error("magnax's scrcpy is incompatible with your PC")
            logger.info("Please install the software yourself : brew install scrcpy")    
        return result
    
    @classmethod
    def stop_record(cls):
        logger.info('stop scrcpy process')
        pids = psutil.pids()
        scrcpy_processes = []

        # 首先找到所有scrcpy进程
        try:
            for pid in pids:
                try:
                    p = psutil.Process(pid)
                    if p.name().__contains__('scrcpy'):
                        scrcpy_processes.append(p)
                        logger.info(f'发现scrcpy进程: {pid}')
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    continue
        except Exception as e:
            logger.exception(e)

        # 尝试温和地终止进程，给scrcpy足够时间完成MKV容器写入
        if scrcpy_processes:
            logger.info('尝试温和地停止scrcpy进程...')
            for process in scrcpy_processes:
                try:
                    process.terminate()  # 发送SIGTERM信号
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    continue

            # 等待进程优雅退出，scrcpy需要时间来正确关闭MKV容器（写入索引和元数据）
            gone, alive = psutil.wait_procs(scrcpy_processes, timeout=15)

            if gone:
                logger.info(f'scrcpy进程已优雅退出: {[p.pid for p in gone]}')

            if alive:
                logger.warning(f'scrcpy进程超时未退出，强制终止: {[p.pid for p in alive]}')
                for process in alive:
                    try:
                        process.kill()
                        logger.info(f'强制终止进程: {process.pid}')
                    except (psutil.NoSuchProcess, psutil.AccessDenied):
                        continue
                time.sleep(1)

            logger.info('scrcpy进程停止完成')
        else:
            logger.info('没有发现运行中的scrcpy进程')
    
    @classmethod
    def cast_screen(cls, device):
        logger.info('start cast screen')
        win_cmd = "start /i {scrcpy_path} -s {deviceId} --no-power-on".format(
            scrcpy_path = cls.scrcpy_path(), 
            deviceId = device
        )
        mac_cmd = "nohup {scrcpy_path} -s {deviceId} --no-power-on &".format(
            scrcpy_path = cls.scrcpy_path(), 
            deviceId = device
        )
        if platform.system().lower().__contains__('windows'):
            result = os.system(win_cmd)
        else:
            result = os.system(mac_cmd)
        if result == 0:
            logger.info("cast screen success")
        else:
            logger.error("magnax's scrcpy is incompatible with your PC")
            logger.info("Please install the software yourself : brew install scrcpy")    
        return result
    
    @classmethod
    def play_video(cls, video):
        logger.info('start play video : {}'.format(video))
        cap = cv2.VideoCapture(video)
        while(cap.isOpened()):
            ret, frame = cap.read()
            if ret:
                gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                cv2.namedWindow("frame", 0)  
                cv2.resizeWindow("frame", 430, 900)
                cv2.imshow('frame', gray)
                if cv2.waitKey(25) & 0xFF == ord('q') or not cv2.getWindowProperty("frame", cv2.WND_PROP_VISIBLE):
                    break
            else:
                break
        cap.release()
        cv2.destroyAllWindows()
